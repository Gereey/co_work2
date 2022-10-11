import matplotlib.pyplot as plt  #画图用
import numpy as np
import matplotlib
import operator
import openpyxl
import random
import xlrd 				#读Excel数据用
import pandas as pd
import openpyxl
from openpyxl import Workbook,load_workbook

def data(filename):
    wb = load_workbook('./list1.xlsx')
    sheet = wb.active
    # 引入文件
    tim = []                                    #经常旷课同学的旷课时间（第几节课）
    y = random.randint(5,8)
    #print(y)
    for i in range(1, y+1):
        l = random.sample(range(1, 21), 16)
        tim.append(l)

    abs_stu = random.sample(range(1, 91), y)    #经常旷课的同学的名单
    '''
    print(" #经常旷课的同学的名单")
    print(abs_stu)
    '''
    abStu = {absstu:time for absstu,time in zip(abs_stu,tim)}
    '''print(" #经常旷课的同学的名单和时间")
    print(abStu)
    '''
    for i in abStu:
        for j in abStu[i]:
            sheet.cell(i+1, j+2).value = 0      #将相应的时间赋值为0


    fre_abs = []
    for i in range(1, 21):
        once_thre_abs = []
        n = random.randint(0,3)
        while len(once_thre_abs) != n:
            unit = random.randint(1, 90)
            if unit not in once_thre_abs and unit not in abs_stu:
                once_thre_abs.append(unit)
        fre_abs.append(once_thre_abs)

    #print(" #20次，偶尔旷课的同学的总次数")
    x = 0
    for i in range(0,len(fre_abs)):
        x += len(fre_abs[i])
    #print(x)
    #print(" #所有旷课的同学的总次数")

    #print(x+y*16)
    for i in range(len(fre_abs)):
        for j in fre_abs[i]:
            sheet.cell(j+1, i+3).value = 0

    for i in abStu:
        for j in abStu[i]:
            sheet.cell(i+1, j+2).value = 0


    for i in range(2,92):
        for j in range(3,23):
            if sheet.cell(i, j).value is None:
                sheet.cell(i, j).value = 1

    wb.save(filename)

if __name__ == "__main__":
    data("list.xlsx")