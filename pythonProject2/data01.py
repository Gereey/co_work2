import matplotlib.pyplot as plt  #画图用
import numpy as np
import matplotlib
import operator
import openpyxl
import random
import xlrd 				#读Excel数据用
import pandas as pd
import openpyxl
import data
from openpyxl import Workbook,load_workbook

def main():
    data.data()
    wb = load_workbook('./list6.xlsx')
    sheet = wb.active

    valid = 0
    total = 0
    al_absence = []
    been_absence = []

    for i in range(3, 23):

        if i == 3:
            for j in range(2,92):
                if sheet.cell(j, i).value == 0:
                    been_absence.append(j)
                    valid+=1
                    #print(j)
                total+=1

        elif i == 4:
            check = al_absence + been_absence
            for j in check:
                if sheet.cell(j, i).value == 0:
                    al_absence.append(j)
                    been_absence.remove(j)
                    valid+=1
                    #print(j)
                total+=1
        elif i == 5:
            check = al_absence + been_absence
            for j in check:
                if sheet.cell(j, i).value == 0:
                    if j not in al_absence:
                        al_absence.append(j)
                        been_absence.remove(j)
                    valid += 1
                else:
                    if j in been_absence:
                        been_absence.remove(j)

                    # print(j)
                total += 1

        else:
            for j in al_absence:
                if sheet.cell(j, i).value == 0:
                    valid += 1
                    # print(j)
                total += 1

        '''else:
            check = al_absence
            for j in check:
                if sheet.cell(j, i).value == 0 :
                    valid += 1
                    # print(j)
                     total += 1
         '''
    print(len(al_absence))
    print(valid)
    print(total)
    print(valid/total)


if __name__ == "__main__":
    main()




