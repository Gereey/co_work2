import matplotlib.pyplot as plt  #画图用
import numpy as np
import matplotlib
import operator
import openpyxl
import random
import xlrd 				#读Excel数据用
import pandas as pd
import openpyxl
import data
from openpyxl import Workbook,load_workbook




def main():
    #data.data()
    wb = load_workbook('./list6.xlsx')
    sheet = wb.active

    valid = 0
    total = 0
    al_absence = []
    been_absence = []

    for i in range(3, 23):

        check = al_absence + been_absence

        if len(check) == 0:
            for j in range(2,92):
                if sheet.cell(j, i).value == 0:
                    been_absence.append(j)
                    valid+=1
                    #print(j)
                total+=1

        elif i == 4:
            for j in check:
                if sheet.cell(j, i).value == 1:
                    been_absence.remove(j)
                else:
                    valid += 1
                    # print(j)
                total += 1

        else:
            for j in check:
                if sheet.cell(j, i).value == 0:
                    valid += 1
                    #print(j)
                total += 1

    print(valid)
    print(total)
    print(valid/total)


if __name__ == "__main__":
    main()










'''   
if len(al_absence) >= 5:
    check1 = random.shuffle(been_absence)
    ckeck = al_absence+check1[0:3]
    for i in check:
        if i in abStu:
           if i in been_absence:
               been_absence[i] = been_absence.get(i) + 1
              # if been_absence.get(i) > 4:
                #   pass
            valid+=1
        total+=1
'''

