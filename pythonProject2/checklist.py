import matplotlib.pyplot as plt  #画图用
import numpy as np
import matplotlib
import operator
import openpyxl
import random
import xlrd 				#读Excel数据用
import pandas as pd
import openpyxl
import data
from openpyxl import Workbook,load_workbook

def main():
    subject = ['sample1.xlsx', 'sample2.xlsx', 'sample3.xlsx', 'sample4.xlsx', 'sample5.xlsx']
    namelist = ['namelist1.xlsx', 'namelist2.xlsx', 'namelist3.xlsx', 'namelist4.xlsx', 'namelist5.xlsx']
    k = 0
    valid = 0
    total = 0
    while k < 5:
        data.data(subject[k])
        wb = load_workbook(subject[k])
        sheet = wb.active
        nl = Workbook()
        desheet = nl.active

        al_absence = []
        been_absence = []

        for i in range(3, 23):
            desheet.cell(1, i-2).value = i-2
            n = 2
            if i == 3:
                for j in range(2, 92):
                    if sheet.cell(j, i).value == 0:
                        been_absence.append(j)
                        valid += 1
                        #print(j)
                    total += 1
                    desheet.cell(n, i-2).value = sheet.cell(j, 2).value
                    n += 1
            elif i == 4:
                check = al_absence + been_absence
                for j in check:
                    if sheet.cell(j, i).value == 0:
                        al_absence.append(j)
                        been_absence.remove(j)
                        valid += 1
                        #print(j)
                    total+=1
                    desheet.cell(n, i - 2).value = sheet.cell(j, 2).value
                    n += 1
            elif i == 5:
                check = al_absence + been_absence
                for j in check:
                    if sheet.cell(j, i).value == 0:
                        if j not in al_absence:
                            al_absence.append(j)
                            been_absence.remove(j)
                        valid += 1
                    else:
                        if j in been_absence:
                            been_absence.remove(j)

                        # print(j)
                    total += 1
                    desheet.cell(n, i - 2).value = sheet.cell(j, 2).value
                    n += 1

            else:
                for j in al_absence:
                    if sheet.cell(j, i).value == 0:
                        valid += 1
                        # print(j)
                    total += 1
                    desheet.cell(n, i - 2).value = sheet.cell(j, 2).value
                    n += 1

        nl.save(namelist[k])
        k += 1
    # print(valid)
    # print(total)
    # print(valid / total)
    #print(len(al_absence))
    return valid/ total



if __name__ == "__main__":
    s = 0
    for i in range(1, 101):
        s += main()
        print(s)
    print(s/100)




